
import sklearn
from openpyxl.reader.excel import load_workbook


def lodData(path, tag, BinaryDecisionVector):
    wb = load_workbook(filename = path)
    sheet_name = wb.get_sheet_names()
    sheet = wb.get_sheet_by_name(sheet_name[0])
    data = []
    for j in range(14):
        for i in range(19):
            data.append(sheet.cell(row= i+1, column= j+1).value)

    # for i in range(19):
    #     data.append(sheet.cell(row=i + 1, column=14).value)

    for i in range(len(BinaryDecisionVector)):
        if BinaryDecisionVector[i] == 0:
            data[i] == 0

    data.append(tag)
    return data
